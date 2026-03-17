import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import psutil
import unicodedata
import threading
import time

def norm(s):
    return unicodedata.normalize('NFC', str(s)) if s is not None else ""

class PreviewWindow(tk.Toplevel):
    def __init__(self, parent, results, main_app, keyword):
        super().__init__(parent)
        self.title("검색 결과 상세 미리보기")
        self.geometry("950x650")
        self.main_app = main_app
        self.results = results
        self.keyword = keyword
        
        total_matches = sum(len(res['details']) for res in results)
        
        info_frame = tk.Frame(self, padx=10, pady=10, bg="#f0f0f0")
        info_frame.pack(fill=tk.X)
        
        tk.Label(info_frame, text=f"검색어: '{keyword}'  |  총 매칭 건수: {total_matches}건", font=("Arial", 10, "bold"), bg="#f0f0f0").pack(side=tk.LEFT, padx=5)
        tk.Label(info_frame, text="   →   변경할 문구:", bg="#f0f0f0").pack(side=tk.LEFT, padx=5)
        self.edit_replace_keyword = tk.Entry(info_frame, width=20)
        self.edit_replace_keyword.insert(0, self.main_app.replace_keyword.get())
        self.edit_replace_keyword.pack(side=tk.LEFT, padx=5)
        
        frame = tk.Frame(self, padx=10, pady=10)
        frame.pack(fill=tk.BOTH, expand=True)
        
        tree_frame = tk.Frame(frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        columns = ("select", "file", "type", "detail")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=18)
        self.tree.heading("select", text="선택")
        self.tree.heading("file", text="파일명")
        self.tree.heading("type", text="구분")
        self.tree.heading("detail", text="상세 내용 (라인/셀)")
        
        self.tree.column("select", width=50, anchor=tk.CENTER)
        self.tree.column("file", width=200)
        self.tree.column("type", width=100)
        self.tree.column("detail", width=550)
        
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.match_map = {}
        for res in results:
            for match in res['details']:
                match['selected'] = True
                # 태그 추가: 삭제 기능에서 구분하기 위해 필요 (Name 또는 Name (Phone) 모두 포함)
                tag = "file_name" if "Name" in match['type'] else "content"
                item_id = self.tree.insert("", tk.END, values=("☑", res['name'], match['type'], match['content']), tags=(tag,))
                # match_map에 전체 정보를 담아둠
                self.match_map[item_id] = {"file_res": res, "match_info": match}
        
        self.tree.match_map = self.match_map # execute_delete_items 에서 인식용 헬퍼
        self.tree.bind("<ButtonRelease-1>", self.on_item_click)
        
        btn_frame = tk.Frame(self, pady=10)
        btn_frame.pack(fill=tk.X)
        
        tk.Button(btn_frame, text="전체 선택", command=self.select_all, width=12).pack(side=tk.LEFT, padx=10)
        tk.Button(btn_frame, text="전체 해제", command=self.deselect_all, width=12).pack(side=tk.LEFT, padx=5)
        
        self.btn_change = tk.Button(btn_frame, text="선택 항목 변경 실행", command=self.run_replace, bg="#4CAF50", fg="white", font=("Arial", 10, "bold"), width=20)
        self.btn_change.pack(side=tk.LEFT, padx=10)

        self.btn_delete = tk.Button(btn_frame, text="선택 항목 일괄 삭제", command=self.confirm_delete, bg="#f44336", fg="white", font=("Arial", 10, "bold"), width=20)
        self.btn_delete.pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="확인 완료 (닫기)", command=self.destroy, width=15).pack(side=tk.RIGHT, padx=10)

    def confirm_delete(self):
        selected_items = [info for info in self.match_map.values() if info["match_info"].get("selected")]
        if not selected_items:
            messagebox.showwarning("경고", "삭제할 항목을 선택해주세요.")
            return
        
        count = len(selected_items)
        if messagebox.askyesno("삭제 확인", f"선택한 {count}개 항목을 삭제하시겠습니까?\n\n- 파일명 매칭: 파일 자체가 삭제됩니다.\n- 내용 매칭: 해당 줄/셀이 삭제됩니다.\n\n이 작업은 되돌릴 수 없습니다."):
            # Treeview 노드 ID를 찾아서 처리하거나 match_map 정보를 직접 전달
            # BatchRenamer의 execute_delete_items가 Treeview 노드를 받으므로 노드 리스트 구성
            selected_nodes = [node_id for node_id, info in self.match_map.items() if info["match_info"].get("selected")]
            self.main_app.execute_delete_items(selected_nodes, self.tree)
            # 삭제 후 트리에서 제거
            for node in selected_nodes:
                try: 
                    self.tree.delete(node)
                    del self.match_map[node]
                except: pass

    def on_item_click(self, event):
        item_id = self.tree.identify_row(event.y)
        if item_id in self.match_map:
            match = self.match_map[item_id]["match_info"]
            match['selected'] = not match['selected']
            self.tree.set(item_id, column="select", value="☑" if match['selected'] else "☐")

    def select_all(self):
        for item_id, info in self.match_map.items():
            info["match_info"]['selected'] = True
            self.tree.set(item_id, column="select", value="☑")

    def deselect_all(self):
        for item_id, info in self.match_map.items():
            info["match_info"]['selected'] = False
            self.tree.set(item_id, column="select", value="☐")

    def run_replace(self):
        selected_matches = [info for info in self.match_map.values() if info["match_info"].get("selected")]
        if not selected_matches:
            messagebox.showwarning("경고", "항목을 선택해 주세요.")
            return
        target = self.keyword
        replace = self.edit_replace_keyword.get() 
        if not messagebox.askyesno("최종 확인", f"{len(selected_matches)}개 항목을 변경하시겠습니까?"):
            return
        self.main_app.execute_specific_replace(selected_matches, target, replace)
        self.destroy()

class BatchRenamer:
    def check_memory_safe(self, silent=False):
        try:
            mem = psutil.virtual_memory()
            if mem.available / mem.total < 0.02:
                if not silent: print(f"[SYSTEM] Critical Memory: {mem.percent}% used.")
                return False
            return True
        except: return True

    def __init__(self, root):
        self.root = root
        self.root.title("파일/폴더 일괄 이름 변경 도구")
        self.root.geometry("950x800")
        self.content_search_var = tk.BooleanVar(value=False)
        self.phone_search_var = tk.BooleanVar(value=False)
        self.list_only_search_var = tk.BooleanVar(value=False)
        self.doc_extensions = ['.xls', '.xlsx', '.xlsm', '.xlsb', '.txt', '.log', '.csv', '.doc', '.docx', '.ppt', '.pptx', '.hwp', '.hwpx', '.ini']
        self.is_working = False
        self.stop_event = threading.Event()
        self.search_results = [] # 누락된 변수 추가
        self.setup_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing) # 윈도우 X 버튼 연동

    def setup_ui(self):
        # 상단 제작자 표시
        author_frame = tk.Frame(self.root, bg="#333333")
        author_frame.pack(fill=tk.X)
        tk.Label(author_frame, text="제작자: SaRaM_ida(망고아빠)", fg="white", bg="#333333", font=("Malgun Gothic", 9)).pack(side=tk.RIGHT, padx=10, pady=2)

        top_frame = tk.Frame(self.root, pady=10)
        top_frame.pack(fill=tk.X)
        tk.Label(top_frame, text="검색 경로:").pack(side=tk.LEFT, padx=5)
        self.path_entry = tk.Entry(top_frame, width=50)
        self.path_entry.insert(0, "D:/")
        self.path_entry.pack(side=tk.LEFT, padx=5)
        tk.Button(top_frame, text="경로 선택", command=self.browse_path).pack(side=tk.LEFT, padx=5)
        tk.Button(top_frame, text="파일 선택", command=self.browse_files).pack(side=tk.LEFT, padx=5)

        input_frame = tk.Frame(self.root, pady=10)
        input_frame.pack(fill=tk.X)
        
        # Grid 레이아웃 재조정
        tk.Label(input_frame, text="기존 이름:").grid(row=0, column=0, padx=5, sticky="e")
        self.search_keyword = tk.Entry(input_frame, width=15)
        self.search_keyword.grid(row=0, column=1, padx=5)
        
        tk.Label(input_frame, text="변경 이름:").grid(row=0, column=2, padx=5, sticky="e")
        self.replace_keyword = tk.Entry(input_frame, width=15)
        self.replace_keyword.grid(row=0, column=3, padx=5)

        self.btn_search = tk.Button(input_frame, text="검색하기", command=self.start_search_thread, bg="#4CAF50", fg="white", width=12)
        self.btn_search.grid(row=0, column=4, padx=5)
        
        tk.Button(input_frame, text="사용 설명서(?)", command=self.show_help, bg="#2196F3", fg="white", width=12).grid(row=0, column=5, padx=5)
        tk.Button(input_frame, text="프로그램 종료", command=self.on_closing, bg="#757575", fg="white", width=12).grid(row=0, column=6, padx=5)

        self.content_check = tk.Checkbutton(input_frame, text="파일 내용 검색 포함", variable=self.content_search_var)
        self.content_check.grid(row=1, column=0, columnspan=2, sticky=tk.W, padx=5)
        self.phone_check = tk.Checkbutton(input_frame, text="전화번호 검색", variable=self.phone_search_var)
        self.phone_check.grid(row=1, column=2, sticky=tk.W, padx=5)
        self.list_only_check = tk.Checkbutton(input_frame, text="현재 리스트 내에서 검색", variable=self.list_only_search_var)
        self.list_only_check.grid(row=1, column=3, columnspan=2, sticky=tk.W, padx=5)

        list_frame = tk.Frame(self.root)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10)
        columns = ("select", "type", "match", "name", "path")
        self.tree = ttk.Treeview(list_frame, columns=columns, show="headings")
        for col in columns: self.tree.heading(col, text=col.capitalize())
        self.tree.column("select", width=50); self.tree.column("path", width=500)
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.bind("<ButtonRelease-1>", self.on_item_click)

        progress_frame = tk.Frame(self.root, pady=10)
        progress_frame.pack(fill=tk.X, padx=10)
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill=tk.X, side=tk.TOP, pady=5)
        
        status_btn_frame = tk.Frame(progress_frame)
        status_btn_frame.pack(fill=tk.X)
        self.status_label = tk.Label(status_btn_frame, text="준비 완료", fg="blue", font=("Arial", 9))
        self.status_label.pack(side=tk.LEFT)
        
        self.btn_stop = tk.Button(status_btn_frame, text="중단 (Stop)", command=self.stop_work, bg="#FF9800", fg="black", font=("Arial", 9, "bold"), state=tk.DISABLED)
        self.btn_stop.pack(side=tk.RIGHT, padx=5)

        bottom_frame = tk.Frame(self.root, pady=10)
        bottom_frame.pack(fill=tk.X)
        tk.Button(bottom_frame, text="전체 선택", command=self.select_all).pack(side=tk.LEFT, padx=10)
        tk.Button(bottom_frame, text="전체 해제", command=self.deselect_all).pack(side=tk.LEFT, padx=5)
        
        self.btn_main_delete = tk.Button(bottom_frame, text="선택 항목 일괄 삭제", command=self.confirm_main_delete, bg="#f44336", fg="white", font=("Arial", 10, "bold"), width=20)
        self.btn_main_delete.pack(side=tk.RIGHT, padx=10)

        self.btn_rename = tk.Button(bottom_frame, text="일괄 이름 변경 실행", command=self.start_rename_thread, bg="#4CAF50", fg="white", font=("Arial", 10, "bold"), width=20)
        self.btn_rename.pack(side=tk.RIGHT, padx=5)

    def on_closing(self):
        if self.is_working:
            if messagebox.askyesno("종료 확인", "현재 작업(검색/변경/삭제)이 진행 중입니다.\n작업을 중단하고 프로그램을 종료하시겠습니까?"):
                self.stop_event.set()
                self.root.destroy()
        else:
            if messagebox.askyesno("종료", "프로그램을 종료하시겠습니까?"):
                self.root.destroy()

    def show_help(self):
        help_win = tk.Toplevel(self.root)
        help_win.title("사용 설명서 - batch_renamer_v2")
        help_win.geometry("600x550")
        
        txt = tk.Text(help_win, padx=10, pady=10, font=("Malgun Gothic", 10), wrap=tk.WORD)
        txt.pack(fill=tk.BOTH, expand=True)
        
        help_content = """[ batch_renamer_v2 사용 설명서 ]

1. 검색 기능
- 검색 경로: 작업을 수행할 드라이브나 폴더를 지정합니다.
- 파일 내용 검색 포함:체크 시 텍스트 및 엑셀 파일 내부의 내용까지 검색합니다.
- 현재 리스트 내에서 검색: 이미 검색된 결과 집합 안에서 재검색할 때 사용합니다.

2. 상세 미리보기 및 정밀 변경
- 검색 완료 후 나타나는 팝업창에서 변경을 원하는 특정 라인이나 셀만 선택(☑)할 수 있습니다.
- 상단의 '변경할 문구'를 입력하고 '선택 항목 변경 실행'을 누르면 선택된 부분만 정밀하게 수정됩니다.

3. 중단(Stop) 기능
- 대용량 작업이 너무 오래 걸릴 경우 하단의 '중단(Stop)' 버튼을 눌러 작업을 안전하게 멈출 수 있습니다.

4. 일괄 이름 변경
- 메인 화면 하단의 '일괄 이름 변경 실행' 버튼은 리스트에서 선택된 모든 항목의 파일 이름을 변경합니다.

5. 결과 확인
- 모든 변경 작업이 끝나면 '재검색 여부'를 묻는 팝업이 뜹니다.

제작자: SaRaM_ida(망고아빠)
"""
        txt.insert(tk.END, help_content)
        txt.config(state=tk.DISABLED)
        tk.Button(help_win, text="닫기", command=help_win.destroy, width=10).pack(pady=10)

    def set_working(self, working):
        self.is_working = working
        state = tk.DISABLED if working else tk.NORMAL
        self.btn_search.config(state=state)
        self.btn_rename.config(state=state)
        self.btn_stop.config(state=tk.NORMAL if working else tk.DISABLED)
        if not working:
            self.progress_var.set(0)
            self.status_label.config(text="대기 중")

    def stop_work(self):
        if self.is_working:
            if messagebox.askyesno("확인", "현재 진행 중인 작업을 중단하시겠습니까?"):
                self.stop_event.set()
                self.status_label.config(text="중단 신호 전송됨. 잠시만 기다려주세요...", fg="red")

    def update_progress(self, val, msg):
        self.progress_var.set(val)
        self.status_label.config(text=msg)
        self.root.update_idletasks()

    def start_search_thread(self):
        self.stop_event.clear()
        threading.Thread(target=self.search_items, daemon=True).start()

    def start_rename_thread(self):
        self.stop_event.clear()
        threading.Thread(target=self.rename_items, daemon=True).start()

    def browse_path(self):
        path = filedialog.askdirectory()
        if path: self.path_entry.delete(0, tk.END); self.path_entry.insert(0, path)

    def browse_files(self):
        files = filedialog.askopenfilenames(title="파일 선택", filetypes=[("Excel", "*.xlsx;*.xlsm;*.xlsb;*.xls"), ("Text", "*.txt;*.log;*.csv"), ("All", "*.*")])
        if files:
            for f in files:
                f = os.path.normpath(f)
                if not any(res["path"] == f for res in self.search_results):
                    self.add_to_list("File", "Selected", os.path.basename(f), f)

    def confirm_main_delete(self):
        # 메인 트리에서 체크된 항목들 (현재 tree의 match_map이 없으므로 search_results 기반)
        selected_indices = [i for i, res in enumerate(self.search_results) if res.get("selected")]
        if not selected_indices:
            messagebox.showwarning("경고", "삭제할 항목을 선택해주세요.")
            return
        
        count = len(selected_indices)
        if messagebox.askyesno("삭제 확인", f"메인 리스트에서 선택한 {count}개 항목을 일괄 삭제하시겠습니까?\n\n- 파일명 매칭: 파일 자체가 삭제됩니다.\n- 내용 매칭: 해당 파일의 모든 매칭된 줄/셀이 삭제됩니다.\n\n이 작업은 되돌릴 수 없습니다."):
            # execute_delete_items에 전달할 노드 리스트 구성
            selected_nodes = []
            for i in selected_indices:
                selected_nodes.append(self.search_results[i]["id"])
            
            self.execute_delete_items(selected_nodes, self.tree)
            
            # 삭제 후 검색 결과에서 동기화 (필요 시 재검색 유도)
            if messagebox.askyesno("완료", "삭제 작업이 완료되었습니다.\n리스트를 갱신하시겠습니까?"):
                self.search_items()

    def apply_changes(self):
        base_path = self.path_entry.get(); keyword = self.search_keyword.get(); list_only = self.list_only_search_var.get()
        phone_search = self.phone_search_var.get()
        if not keyword and not phone_search: messagebox.showwarning("경고", "키워드를 입력하세요."); return
        self.set_working(True); self.update_progress(0, "검색 준비 중...")
        
    def search_items(self):
        base_path = self.path_entry.get(); keyword = self.search_keyword.get(); list_only = self.list_only_search_var.get()
        phone_search = self.phone_search_var.get()
        if not keyword and not phone_search: messagebox.showwarning("경고", "키워드를 입력하세요."); return
        self.set_working(True); self.update_progress(0, "검색 준비 중...")
        
        phone_pattern = re.compile(r'010[- ]?\d{3,4}[- ]?\d{4}')
        n_keyword = norm(keyword).strip()
        preview_data = []
        if list_only:
            total = len(self.search_results)
            for i, res in enumerate(self.search_results, 1):
                if self.stop_event.is_set(): break
                self.update_progress((i/total)*100, f"검색 중: {os.path.basename(res['path'])} ({i}/{total})")
                matches = []; match_type = "Selected"
                f_name = os.path.basename(res['path']); ext = os.path.splitext(f_name)[1].lower()
                n_fname = norm(f_name)
                # 파일명 검색 (일반 키워드 또는 전화번호)
                if n_keyword and n_keyword in n_fname:
                    matches.append({"type": "Name", "content": f_name}); match_type = "Name"
                elif phone_search:
                    pm = phone_pattern.search(n_fname)
                    if pm:
                        matches.append({"type": "Name (Phone)", "content": f_name, "raw": pm.group()})
                        match_type = "Name (Phone)"
                if self.content_search_var.get():
                    if phone_search and ext not in self.doc_extensions:
                        matches = []
                    else:
                        c_matches = self._dispatch_content_search(res['path'], ext, keyword, phone_search)
                        if c_matches: matches.extend(c_matches); match_type = self._get_match_label(ext)
                if matches:
                    self.tree.set(res["id"], column="match", value=match_type); res["match"] = match_type
                    preview_data.append({"name": f_name, "path": res['path'], "details": matches})
            
            stopped = self.stop_event.is_set()
            self.set_working(False)
            if stopped: messagebox.showinfo("중단", "사용자에 의해 검색이 중단되었습니다."); return
            if preview_data: PreviewWindow(self.root, preview_data, self, keyword)
            else: messagebox.showinfo("알림", "검색 결과 없음")
        else:
            for item in self.tree.get_children(): self.tree.delete(item)
            self.search_results = []
            try:
                all_files = []
                for root, _, files in os.walk(base_path):
                    if self.stop_event.is_set(): break
                    for f in files: all_files.append(os.path.join(root, f))
                
                total = len(all_files)
                for i, full_path in enumerate(all_files, 1):
                    if self.stop_event.is_set(): break
                    if i % 10 == 0: self.update_progress((i/total)*100, f"검색 중... ({i}/{total})")
                    f = os.path.basename(full_path); ext = os.path.splitext(f)[1].lower(); m_type = ""
                    n_fname = norm(f)
                    item_matches = []
                    # 파일명 검색 (일반 키워드 또는 전화번호)
                    if n_keyword and n_keyword in n_fname:
                        m_type = "Name"; item_matches.append({"type": "Name", "content": f})
                    elif phone_search:
                        pm = phone_pattern.search(n_fname)
                        if pm:
                            m_type = "Name (Phone)"
                            item_matches.append({"type": "Name (Phone)", "content": f, "raw": pm.group()})
                    if self.content_search_var.get():
                        if phone_search and ext not in self.doc_extensions:
                            c_matches = []
                        else:
                            c_matches = self._dispatch_content_search(full_path, ext, keyword, phone_search)
                        if c_matches: m_type = self._get_match_label(ext); item_matches.extend(c_matches)
                    if m_type: 
                        self.add_to_list("File", m_type, f, full_path)
                        preview_data.append({"name": f, "path": full_path, "details": item_matches})
                
                stopped = self.stop_event.is_set()
                self.set_working(False)
                if stopped: messagebox.showinfo("중단", "사용자에 의해 검색이 중단되었습니다."); return
                if not self.search_results: messagebox.showinfo("알림", "검색 결과 없음")
                else: PreviewWindow(self.root, preview_data, self, keyword)
            except Exception as e: self.set_working(False); messagebox.showerror("오류", str(e))

    def _dispatch_content_search(self, path, ext, keyword, phone_search=False):
        if self.stop_event.is_set(): return []
        if ext in ['.txt', '.log', '.csv', '.ini']: return self.search_in_text(path, keyword, phone_search)
        if ext in ['.xlsx', '.xlsm']: return self.search_in_excel(path, keyword, phone_search)
        if ext in ['.xlsb']: return self.search_in_xlsb(path, keyword, phone_search)
        if ext in ['.xls']: return self.search_in_xls(path, keyword, phone_search)
        return []

    def _get_match_label(self, ext): return "Content (Excel)" if ext in ['.xlsx', '.xlsm', '.xlsb', '.xls'] else "Content (TXT)"

    def search_in_text(self, file_path, keyword, phone_search=False):
        matches = []; n_keyword = norm(keyword)
        phone_pattern = re.compile(r'010[- ]?\d{3,4}[- ]?\d{4}')
        for enc in ['utf-8', 'cp949', 'euc-kr']:
            if self.stop_event.is_set(): break
            try:
                with open(file_path, 'r', encoding=enc) as f:
                    for i, line in enumerate(f, 1):
                        if self.stop_event.is_set(): break
                        n_line = norm(line)
                        if phone_search:
                            found = phone_pattern.findall(n_line)
                            for p in found:
                                matches.append({"type": f"Line {i} (Phone)", "content": f"{p} (in: {line.strip()[:50]}...)", "pos": i, "raw": p})
                        if n_keyword and n_keyword in n_line:
                            matches.append({"type": f"Line {i}", "content": line.strip()[:100], "pos": i})
                return matches
            except: continue
        return []

    def search_in_excel(self, file_path, keyword, phone_search=False):
        import openpyxl
        n_keyword = norm(keyword); matches = []
        phone_pattern = re.compile(r'010[- ]?\d{3,4}[- ]?\d{4}')
        try:
            for data_only in [True, False]:
                if self.stop_event.is_set(): break
                wb = openpyxl.load_workbook(file_path, data_only=data_only, read_only=True)
                for sheet in wb.worksheets:
                    if self.stop_event.is_set(): break
                    for r, row in enumerate(sheet.iter_rows(values_only=True), 1):
                        if self.stop_event.is_set(): break
                        for c, val in enumerate(row, 1):
                            if self.stop_event.is_set(): break
                            if val is None: continue
                            s_val = norm(val)
                            if phone_search:
                                found = phone_pattern.findall(s_val)
                                for p in found:
                                    matches.append({"type": "Cell (Phone)", "content": f"[{sheet.title}!{openpyxl.utils.get_column_letter(c)}{r}] {p}", "sheet": sheet.title, "row": r, "col": c, "raw": p})
                                # 인접 3개 셀 결합 체크 (010 | 1234 | 5678 형식 지원)
                                if c <= len(row) - 2:
                                    combined = "".join(str(x).strip() if x is not None else "" for x in row[c-1:c+2])
                                    if phone_pattern.fullmatch(combined):
                                        matches.append({"type": "Cell (Phone/Split)", "content": f"[{sheet.title}!{openpyxl.utils.get_column_letter(c)}{r}~] {combined}", "sheet": sheet.title, "row": r, "col": c, "raw": combined})
                            if n_keyword and n_keyword in s_val:
                                matches.append({"type": "Cell", "content": f"[{sheet.title}!{openpyxl.utils.get_column_letter(c)}{r}] {str(val)[:100]}", "sheet": sheet.title, "row": r, "col": c})
                wb.close()
                if matches: break
        except: pass
        return matches

    def search_in_xls(self, file_path, keyword, phone_search=False):
        import xlrd
        n_keyword = norm(keyword); matches = []
        phone_pattern = re.compile(r'010[- ]?\d{3,4}[- ]?\d{4}')
        try:
            wb = xlrd.open_workbook(file_path)
            for s in wb.sheets():
                if self.stop_event.is_set(): break
                for r in range(s.nrows):
                    if self.stop_event.is_set(): break
                    for c in range(s.ncols):
                        if self.stop_event.is_set(): break
                        val = s.cell_value(r, c)
                        if val is None: continue
                        s_val = norm(val)
                        if phone_search:
                            found = phone_pattern.findall(s_val)
                            for p in found:
                                matches.append({"type": "Cell (Phone)", "content": f"[{s.name}!{r+1},{c+1}] {p}", "sheet": s.name, "row": r+1, "col": c+1, "raw": p})
                            # 인접 3개 셀 결합 체크 (xls)
                            if c < s.ncols - 2:
                                val2 = s.cell_value(r, c+1); val3 = s.cell_value(r, c+2)
                                combined = "".join(str(x).strip() for x in [val, val2, val3] if x is not None)
                                if phone_pattern.fullmatch(combined):
                                    matches.append({"type": "Cell (Phone/Split)", "content": f"[{s.name}!{r+1},{c+1}~] {combined}", "sheet": s.name, "row": r+1, "col": c+1, "raw": combined})
                        if n_keyword and n_keyword in s_val: # n_keyword 사용
                            matches.append({"type": "Cell", "content": f"[{s.name}!{r+1},{c+1}] {str(val)[:100]}", "sheet": s.name, "row": r+1, "col": c+1})
        except: pass
        return matches

        n_keyword = norm(keyword); matches = []
        phone_pattern = re.compile(r'010-\d{3,4}-\d{4}')
        try:
            with open_workflow(file_path) as wb:
                for s in wb.sheets:
                    if self.stop_event.is_set(): break
                    with wb.get_sheet(s) as sheet:
                        for row in sheet.rows():
                            if self.stop_event.is_set(): break
                            for cell in row:
                                if cell.v is None: continue
                                s_val = norm(cell.v)
                                if phone_search:
                                    found = phone_pattern.findall(s_val)
                                    for p in found:
                                        matches.append({"type": "Cell (Phone)", "content": f"[{s}!{cell.address}] {p}", "sheet": s, "address": cell.address, "raw": p})
                                    # xlsb의 경우 행 단위 데이터 접근이 구조상 다르므로, 필요 시 별도 구현 가능하나 현재는 단일 셀 위주
                                if n_keyword and n_keyword in s_val:
                                    matches.append({"type": "Cell", "content": f"[{s}!{cell.address}] {str(cell.v)[:100]}", "sheet": s, "address": cell.address})
        except: pass
        return matches

    def execute_specific_replace(self, selected_info, target, replace):
        self.stop_event.clear()
        threading.Thread(target=self._exec_specific_replace_worker, args=(selected_info, target, replace), daemon=True).start()

    def _exec_specific_replace_worker(self, selected_info, target, replace):
        self.set_working(True); self.update_progress(0, "항목 변경 준비 중...")
        file_groups = {}
        for info in selected_info:
            p = info["file_res"]["path"]; file_groups.setdefault(p, []).append(info["match_info"])
        total = len(file_groups)
        for i, (path, matches) in enumerate(file_groups.items(), 1):
            if self.stop_event.is_set(): break
            self.update_progress((i/total)*100, f"변경 중: {os.path.basename(path)} ({i}/{total})")
            ext = os.path.splitext(path)[1].lower()
            try:
                # Name 매칭이 하나라도 있으면 파일명 변경 시도
                name_match = next((m for m in matches if "Name" in m["type"]), None)
                if name_match:
                    # 키워드가 있다면 키워드 우선, 없다면 휴대전화 매칭값 사용
                    actual_target = target if target else name_match.get("raw", target)
                    if actual_target:
                        new_name = os.path.basename(path).replace(actual_target, replace)
                        os.rename(path, os.path.join(os.path.dirname(path), new_name))
                    continue
                
                if ext in ['.txt', '.log', '.csv']: self._replace_specific_text(path, matches, target, replace)
                elif ext in ['.xlsx', '.xlsm']: self._replace_specific_excel(path, matches, target, replace)
                else: self.replace_in_xls(path, target, replace)
            except: pass
        
        stopped = self.stop_event.is_set()
        self.set_working(False)
        if stopped: messagebox.showinfo("중단", "작업이 중단되었습니다. 현재까지 완료된 항목만 반영되었습니다."); return
        if messagebox.askyesno("완료", "변경 작업이 완료되었습니다.\n재검색을 통해 변경 결과를 확인하시겠습니까?"):
            self.search_items()

    def _replace_specific_text(self, path, matches, target, replace):
        lines = [m["pos"] for m in matches if "Line" in m["type"]]; tmp = path + ".tmp"
        for enc in ['utf-8', 'cp949', 'euc-kr']:
            try:
                with open(path, 'r', encoding=enc) as f_in, open(tmp, 'w', encoding=enc) as f_out:
                    for i, line in enumerate(f_in, 1): f_out.write(line.replace(target, replace) if i in lines else line)
                os.remove(path); os.rename(tmp, path); return
            except: continue

    def _replace_specific_excel(self, path, matches, target, replace):
        import openpyxl
        wb = openpyxl.load_workbook(path)
        for m in matches:
            if m["type"] == "Cell":
                sheet = wb[m["sheet"]]; cell = sheet.cell(row=m["row"], column=m["col"])
                if cell.value: cell.value = str(cell.value).replace(target, replace)
        wb.save(path)

    def rename_items(self):
        if not self.search_results:
            messagebox.showwarning("경고", "검색을 먼저 진행해 주세요.")
            return
        target = self.search_keyword.get(); replace = self.replace_keyword.get(); selected = [res for res in self.search_results if res["selected"]]
        if not selected: messagebox.showwarning("경고", "항목을 선택하세요."); return
        if not messagebox.askyesno("확인", f"{len(selected)}개를 변경하시겠습니까?"): return
        self.set_working(True); success = fail = 0; selected.sort(key=lambda x: len(x["path"]), reverse=True)
        for i, res in enumerate(selected, 1):
            if self.stop_event.is_set(): break
            self.update_progress((i/len(selected))*100, f"처리 중: {res['name']} ({i}/{len(selected)})")
            try:
                if res["match"] == "Name":
                    os.rename(res["path"], os.path.join(os.path.dirname(res["path"]), res["name"].replace(target, replace)))
                else:
                    ext = os.path.splitext(res["path"])[1].lower()
                    if ext in ['.txt', '.log', '.csv']: self.replace_in_text(res["path"], target, replace)
                    elif ext in ['.xlsx', '.xlsm']: self.replace_in_excel(res["path"], target, replace)
                    else: self.replace_in_xls(res["path"], target, replace)
                success += 1
            except: fail += 1
        
        stopped = self.stop_event.is_set()
        self.set_working(False)
        if stopped: messagebox.showinfo("중단", f"작업이 사용자에 의해 중단되었습니다.\n(현재까지 성공: {success}, 실패: {fail})"); return
        if messagebox.askyesno("완료", f"성공: {success}, 실패: {fail}\n재검색을 통해 변경 결과를 확인하시겠습니까?"):
            self.search_items()

    def replace_in_text(self, path, t, r):
        tmp = path + ".tmp"
        for enc in ['utf-8', 'cp949', 'euc-kr']:
            try:
                with open(path, 'r', encoding=enc) as f_in, open(tmp, 'w', encoding=enc) as f_out:
                    for line in f_in: f_out.write(line.replace(t, r))
                os.remove(path); os.rename(tmp, path); return
            except: continue

    def replace_in_excel(self, path, t, r):
        import openpyxl
        wb = openpyxl.load_workbook(path)
        for s in wb.worksheets:
            for row in s.iter_rows():
                for cell in row:
                    if cell.value: cell.value = str(cell.value).replace(t, r)
        wb.save(path)

    def replace_in_xls(self, path, t, r):
        import win32com.client
        try:
            ex = win32com.client.Dispatch("Excel.Application")
            ex.Visible = False; ex.DisplayAlerts = False
            wb = ex.Workbooks.Open(os.path.abspath(path))
            for s in wb.Sheets: s.Cells.Replace(What=t, Replacement=r)
            wb.Save(); wb.Close(); ex.Quit()
        except: pass

    def add_to_list(self, t, m, n, p):
        item_id = self.tree.insert("", tk.END, values=("☐", t, m, n, p))
        self.search_results.append({"id": item_id, "selected": False, "type": t, "match": m, "name": n, "path": p})

    def on_item_click(self, event):
        item_id = self.tree.identify_row(event.y)
        if not item_id: return
        for res in self.search_results:
            if res["id"] == item_id:
                res["selected"] = not res["selected"]
                self.tree.set(item_id, column="select", value="☑" if res["selected"] else "☐"); break

    def select_all(self):
        for res in self.search_results: res["selected"] = True; self.tree.set(res["id"], column="select", value="☑")

    def deselect_all(self):
        for res in self.search_results: res["selected"] = False; self.tree.set(res["id"], column="select", value="☐")


    def execute_delete_items(self, selected_nodes, tree):
        # PreviewWindow의 match_map을 활용하거나 tree 노드에서 정보를 가져옴
        content_tasks = {} 
        file_to_delete = set()
        
        # 중요: PreviewWindow의 경우 tree가 treeview 객체임
        # tree.match_map이 있으면 PreviewWindow에서 호출된 것임
        is_preview = hasattr(tree, "match_map")
        
        for node in selected_nodes:
            tags = tree.item(node, "tags")
            if is_preview:
                # PreviewWindow 전용 로직
                info = tree.match_map[node]
                path = info["file_res"]["path"]
                if "file_name" in tags:
                    file_to_delete.add(path)
                else:
                    item_info = tree.item(node, "values") # [select, file, type, detail]
                    content_tasks.setdefault(path, []).append({
                        "type": item_info[2],
                        "info": item_info,
                        "match_info": info["match_info"] # 원본 매칭 정보 추가
                    })
            else:
                # 메인 윈도우 전용 로직
                values = tree.item(node, "values") # ("☑"/"☐", Type, Match, Name, Path)
                path = values[4]
                m_type = values[2]
                if m_type == "Name" or m_type == "Name (Phone)":
                    file_to_delete.add(path)
                else:
                    # 메인 화면에서 내용 삭제 시, 해당 파일 내의 모든 매칭을 가져옴
                    # self.search_results에서 해당 파일의 details를 가져와서 tasks로 구성
                    res = next((r for r in self.search_results if r["path"] == path), None)
                    if res and "details" in res:
                        for m in res["details"]:
                            if m["type"] != "Name": # 파일명 매칭 제외한 내용 매칭들
                                content_tasks.setdefault(path, []).append({
                                    "type": m["type"],
                                    "info": [None, None, m["type"], m["content"]], # 헬퍼 대응 형식
                                    "match_info": m
                                })

        # 1. 파일 자체 삭제
        success_files = 0; fail_files = 0
        for f_path in file_to_delete:
            try:
                if os.path.exists(f_path):
                    os.remove(f_path)
                    success_files += 1
            except Exception as e:
                fail_files += 1
                print(f"File delete fail: {f_path}, {e}")

        # 2. 내용 삭제 (파일별 처리)
        success_content = 0; fail_content = 0
        for f_path, tasks in content_tasks.items():
            if f_path in file_to_delete: continue # 이미 삭제된 파일은 건너뛰기
            
            try:
                ext = os.path.splitext(f_path)[1].lower()
                if ext in ['.xls', '.xlsx', '.xlsm', '.xlsb']:
                    # 엑셀 처리
                    self._delete_excel_content(f_path, tasks)
                    success_content += len(tasks)
                else:
                    # 텍스트 처리 (라인 삭제)
                    self._delete_text_content(f_path, tasks)
                    success_content += len(tasks)
            except Exception as e:
                fail_content += len(tasks)
                print(f"Content delete fail: {f_path}, {e}")

        total_msg = f"삭제 완료\n\n- 파일 삭제: {success_files}건 성공"
        if fail_files: total_msg += f" ({fail_files}건 실패)"
        total_msg += f"\n- 내용 삭제: {success_content}건 성공"
        if fail_content: total_msg += f" ({fail_content}건 실패)"
        messagebox.showinfo("결과", total_msg)

    def _delete_text_content(self, file_path, tasks):
        lines_to_remove = set()
        for t in tasks:
            if "match_info" in t and "pos" in t["match_info"]:
                lines_to_remove.add(int(t["match_info"]["pos"]) - 1)
            else:
                # 레거시 백업: 문자열 파싱
                c_info = t["info"][3] if len(t["info"]) == 4 else t["info"][1]
                import re
                m = re.search(r"Line (\d+)", str(t["info"][2]) + c_info) # 'Type' 또는 'Detail'에서 찾음
                if m: lines_to_remove.add(int(m.group(1)) - 1)
        
        if not lines_to_remove: return
        
        for enc in ['utf-8', 'cp949', 'euc-kr']:
            try:
                with open(file_path, 'r', encoding=enc) as f:
                    lines = f.readlines()
                new_lines = [l for i, l in enumerate(lines) if i not in lines_to_remove]
                with open(file_path, 'w', encoding=enc) as f:
                    f.writelines(new_lines)
                return
            except: continue

    def _delete_excel_content(self, file_path, tasks):
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.xlsx' or ext == '.xlsm':
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            for t in tasks:
                if "match_info" in t and "row" in t["match_info"]:
                    m = t["match_info"]
                    sh_name = m.get("sheet"); row_num = m.get("row"); col_num = m.get("col")
                    if sh_name in wb.sheetnames:
                        ws = wb[sh_name]
                        if "Split" in t["type"]:
                            for i in range(3): ws.cell(row=row_num, column=col_num + i).value = None
                        else:
                            ws.cell(row=row_num, column=col_num).value = None
                else:
                    # 레거시 백업: 문자열 파싱
                    c_info = t["info"][3] if len(t["info"]) == 4 else t["info"][1]
                    import re
                    m = re.match(r"\[(.+)!([A-Z]+)(\d+)\]", c_info)
                    if m:
                        sh_name, col_let, row_num = m.groups()
                        if sh_name in wb.sheetnames:
                            ws = wb[sh_name]; ws[f"{col_let}{row_num}"] = None
                    else:
                        m = re.match(r"\[(.+)!([A-Z]+)(\d+)~\]", c_info)
                        if m:
                            sh_name, col_let, row_num = m.groups()
                            if sh_name in wb.sheetnames:
                                ws = wb[sh_name]
                                c_idx = openpyxl.utils.column_index_from_string(col_let)
                                for i in range(3): ws.cell(row=int(row_num), column=c_idx+i).value = None
            wb.save(file_path)
        # .xls 등은 현재 라이브러리 제약상 수정을 위해 xlutils 등이 추가로 필요하므로 보류하거나 알림

if __name__ == "__main__":
    root = tk.Tk(); app = BatchRenamer(root); root.mainloop()
